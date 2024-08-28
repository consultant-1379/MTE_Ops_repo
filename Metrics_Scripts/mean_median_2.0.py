#!/usr/bin/python
import openpyxl
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)
from datetime import datetime,timedelta, time
import csv
import urllib
import sys
from pyexcel.cookbook import merge_all_to_a_book
import glob
import subprocess
from optparse import OptionParser

parser = OptionParser()
arguments=sys.argv
group_list,ISO_list,both_list=[],[],[]

def group():
        if "--group" in arguments:
                global group_list
                group_list=arguments[3:]
                for i in range(0,len(group_list)):
                        group_list[i]=int(group_list[i])
			

def ISO():
        if "--ISO" in arguments:
		global ISO_list
                ISO_list=arguments[3:]


def both():
        if "--both" in arguments:
		global both_list
                both_list=arguments[3:]
                for item in both_list:
                        if "." in item:
                                ISO_list.append(item)
                        else:
                                group_list.append(int(item))


parser.add_option("--ISO",choices=ISO(), action="store_false", dest="filename",help="Enter ISOs to exclude")
parser.add_option("--group", choices=group(), action="store_false", dest="verbose", default=True,help="Enter Groups to exclude")
parser.add_option("--both", choices=both(), action="store_false", dest="filename",help="Enter ISOs and groups to exculde")
(options, args) = parser.parse_args()

queue_link="https://cifwk-oss.lmera.ericsson.se/ENM/queue/"+arguments[1]+"/csv"
csv_file_name="ENM_"+arguments[1]+"_queue.csv"
xlsx_file_name="ENM_"+arguments[1]+"_queue.xlsx"

urllib.urlretrieve(queue_link,csv_file_name)
merge_all_to_a_book(glob.glob(csv_file_name), xlsx_file_name)

wb = openpyxl.Workbook()
wb=openpyxl.load_workbook(xlsx_file_name)
sheets=wb.get_sheet_names()
ws = wb.get_sheet_by_name(name = sheets[0])
row_count=ws.max_row
total = datetime.now()
cell_A,cell_B,cell_C,cell_D,cell_E,cell_I,cell_K="A","B","C","D","E","I","K"
delta_total,ISO_total=total-total,total-total
count=0
ISO_list,ISO_average,median_list,total_median_list=[],[],[],[]


for i in range(2,row_count+1):
        if ws[cell_B+str(i)].value=="Delivered" and  ws[cell_I+str(i)].value!=0 and ws[cell_A+str(i)].value not in group_list:
                ISO_list.append(str(ws[cell_C+str(i)].value))
existing_ISO_list=list(set(ISO_list))
if "None" in existing_ISO_list:
	existing_ISO_list.remove("None")


for ISO in existing_ISO_list:
	for i in range(2,row_count+1):
		if ws[cell_B+str(i)].value=="Delivered" and  ws[cell_I+str(i)].value!=0 and ws[cell_A+str(i)].value not in group_list and ws[cell_C+str(i)].value==ISO and  "flagged this delivery group as having missing dependencies" not in ws[cell_K+str(i)].value:
			created_date=ws[cell_D+str(i)].value
			delivered_date=ws[cell_E+str(i)].value
			delta=delivered_date-created_date
			median_list.append(delta)
                        delta_total=delta_total+delta
                        count+=1
	if count != 0:
		length=len(sorted(median_list))
        	ISO_total=ISO_total+(delta_total/count)
	        delta_total=total-total
        #count=0
		if length%2==0:
			total_median_list.append((sorted(median_list)[length/2]+ sorted(median_list)[(length/2)-1])/2)
		else:
			total_median_list.append(sorted(median_list)[length/2])
	count = 0
	median_list=[]


total_median_list=sorted(total_median_list)
median_length=len(total_median_list)


print "***************************************************************"
if median_length%2==0:
	print "Median of groups delivered in sprint",arguments[1],"=",((total_median_list[median_length/2]+total_median_list[(median_length/2)-1])/2)
else:
	print "Median of groups delivered in sprint",arguments[1],"=",total_median_list[(median_length/2)]
ISO_average=ISO_total/len(existing_ISO_list)
print "***************************************************************"
print "Average of groups delivered in sprint",arguments[1],"=",ISO_average
print "***************************************************************"
